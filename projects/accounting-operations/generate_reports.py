#!/usr/bin/env python3
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, List, Optional, Tuple
import os

from openpyxl import load_workbook


CODE_ROOT = Path(__file__).resolve().parent
DATA_ROOT = Path(os.environ.get("DATA_ROOT", "/Users/m/Documents/accounting/bank")).resolve()

# Accounts metadata (loaded from config/accounts.toml in CODE_ROOT)
ALIAS_MAP = {}
TYPE_MAP = {}
CURRENCY_MAP = {}
STATUS_MAP = {}

@dataclass
class PeriodRange:
    path: Path
    account_id: str
    account_label: str
    start: date
    end: date
    dest: Optional[Path] = None


def parse_account(folder: str) -> Tuple[str, str]:
    """Split folder into account id and label."""
    m = re.match(r"^(\d+)\s*(.*)$", folder)
    if m:
        acc_id = m.group(1)
        label = m.group(2).strip() or folder
        return acc_id, label
    return folder, folder


def display_name(acc_id: str, label: str) -> str:
    alias = ALIAS_MAP.get(acc_id)
    return alias if alias else label


def is_business(acc_id: str) -> bool:
    return TYPE_MAP.get(acc_id, "").lower() == "business"


def parse_period_text(text: str) -> Optional[Tuple[date, date]]:
    """Parse 'dd.mm.yyyy - dd.mm.yyyy' style strings."""
    m = re.search(r"(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})", text)
    if not m:
        return None
    fmt = "%d.%m.%Y"
    try:
        start = datetime.strptime(m.group(1), fmt).date()
        end = datetime.strptime(m.group(2), fmt).date()
    except ValueError:
        return None
    return start, end


def extract_period(path: Path) -> Optional[Tuple[date, date]]:
    """Locate the period row and return the date range."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None
    ws = wb.active

    # Scan top rows for a cell containing "Период" and a sibling date range
    for r in range(1, 15):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        contains_period = any(isinstance(v, str) and "Период" in v for v in row_vals)
        if not contains_period:
            continue
        for v in row_vals:
            if isinstance(v, str):
                rng = parse_period_text(v)
                if rng:
                    return rng
        # If not in the same row, keep looking in next rows for the text
    return None


def extract_account_id(path: Path) -> Optional[str]:
    """Extract account number from header rows."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None
    ws = wb.active
    for r in range(1, 10):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        for v in row_vals:
            if isinstance(v, str):
                s = v.strip()
                if s.isdigit() and len(s) >= 10:
                    return s
    return None


def iter_periods(root: Path) -> List[PeriodRange]:
    periods: List[PeriodRange] = []
    for path in sorted(root.rglob("*.xlsx")):
        if any(part.startswith(".") or part in {"draft", "renamed-statements"} for part in path.parts):
            continue
        account_id, label = parse_account(path.parent.name)
        rng = extract_period(path)
        if not rng:
            continue
        if not account_id.isdigit() or len(account_id) < 6 or account_id not in ALIAS_MAP:
            acc_from_sheet = extract_account_id(path)
            if acc_from_sheet:
                account_id = acc_from_sheet
                label = acc_from_sheet
        if not is_business(account_id):
            continue
        periods.append(
            PeriodRange(
                path=path,
                account_id=account_id,
                account_label=label,
                start=rng[0],
                end=rng[1],
            )
        )
    return periods


def month_bounds(year: int, month: int) -> Tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def intersect(a: Tuple[date, date], b: Tuple[date, date]) -> Optional[Tuple[date, date]]:
    start = max(a[0], b[0])
    end = min(a[1], b[1])
    if start > end:
        return None
    return start, end


def union_intervals(intervals: Iterable[Tuple[date, date]]) -> List[Tuple[date, date]]:
    sorted_ints = sorted(intervals, key=lambda x: x[0])
    merged: List[Tuple[date, date]] = []
    for s, e in sorted_ints:
        if not merged or s > merged[-1][1] + timedelta(days=1):
            merged.append((s, e))
        else:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))
    return merged


def coverage_for_year(periods: List[PeriodRange], year: int):
    by_account = defaultdict(list)
    for p in periods:
        if p.end.year < year or p.start.year > year:
            # only keep portions that intersect the year
            intr = intersect((p.start, p.end), (date(year, 1, 1), date(year, 12, 31)))
            if not intr:
                continue
            by_account[(p.account_id, p.account_label)].append(intr)
        else:
            intr = intersect((p.start, p.end), (date(year, 1, 1), date(year, 12, 31)))
            if intr:
                by_account[(p.account_id, p.account_label)].append(intr)
    return by_account


def month_status(covered: List[Tuple[date, date]], year: int, month: int) -> str:
    if not covered:
        return "🔴"
    month_range = month_bounds(year, month)
    unioned = union_intervals(covered)
    # Check if full month covered
    for s, e in unioned:
        if s <= month_range[0] and e >= month_range[1]:
            return "🟢"
    # Check any overlap
    for s, e in unioned:
        if intersect((s, e), month_range):
            return "🟠"
    return "🔴"


def gaps_for_year(covered: List[Tuple[date, date]], year: int) -> List[Tuple[date, date]]:
    year_range = (date(year, 1, 1), date(year, 12, 31))
    if not covered:
        return [year_range]
    unioned = union_intervals(covered)
    gaps = []
    cursor = year_range[0]
    for s, e in unioned:
        if cursor < s:
            gaps.append((cursor, s - timedelta(days=1)))
        cursor = e + timedelta(days=1)
    if cursor <= year_range[1]:
        gaps.append((cursor, year_range[1]))
    return gaps


def month_end(d: date) -> date:
    if d.month == 12:
        return date(d.year, 12, 31)
    return date(d.year, d.month + 1, 1) - timedelta(days=1)


def add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    day = min(d.day, month_end(date(y, m, 1)).day)
    return date(y, m, day)


def chunk_gap_three_months(gap: Tuple[date, date]) -> List[Tuple[date, date]]:
    """Split a gap into chunks of at most 3 months, aligned to month ends."""
    chunks = []
    start, end = gap
    cursor = start
    while cursor <= end:
        chunk_end = add_months(cursor, 3) - timedelta(days=1)
        if chunk_end > end:
            chunk_end = end
        chunks.append((cursor, chunk_end))
        cursor = chunk_end + timedelta(days=1)
    return chunks


def format_date(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def build_period_markdown(periods: List[PeriodRange]) -> str:
    lines = ["# Period ranges extracted from files", ""]
    for p in periods:
        lines.append(f"- {p.path.as_posix()} — {format_date(p.start)} - {format_date(p.end)}")
    return "\n".join(lines) + "\n"


def build_heatmap_markdown(by_account, year: int) -> str:
    lines = [f"# {year} coverage by account (monthly)", "", "Legend: 🟢 full month covered | 🟠 partial | 🔴 none", ""]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for (acc_id, label), intervals in by_account.items():
        lines.append(f"## {display_name(acc_id, label)}")
        lines.append("| Month | Coverage |")
        lines.append("| --- | --- |")
        for idx, m in enumerate(months, start=1):
            status = month_status(intervals, year, idx)
            lines.append(f"| {m} | {status} |")
        lines.append("")
    return "\n".join(lines)


def build_heatmap_table(by_account, year: int) -> str:
    lines = [f"# {year} coverage table (monthly)", "", "Legend: 🟢 full | 🟠 partial | 🔴 none", ""]
    accounts = sorted(by_account.keys(), key=lambda x: x[0])
    header = ["Month"] + [display_name(acc_id, label) for acc_id, label in accounts]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for idx, m in enumerate(months, start=1):
        row = [m]
        for acc in accounts:
            row.append(month_status(by_account[acc], year, idx))
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines) + "\n"


def build_gap_tasks(by_account, year: int) -> str:
    lines = [f"# {year} coverage gaps to request", ""]
    for (acc_id, label), intervals in by_account.items():
        gaps = gaps_for_year(intervals, year)
        if not gaps:
            continue
        lines.append(f"- Account {display_name(acc_id, label)}")
        for g in gaps:
            lines.append(f"  - Request: {format_date(g[0])}–{format_date(g[1])}")
            sub = chunk_gap_three_months(g)
            if len(sub) > 1:
                for idx, s in enumerate(sub, start=1):
                    lines.append(f"    - Chunk {idx}: {format_date(s[0])}–{format_date(s[1])}")
        lines.append("")
    content = "\n".join(lines).strip()
    if content.endswith("# {year} coverage gaps to request".format(year=year)):
        content += "\nNo gaps for business accounts."
    return content + "\n"


def build_2023_quarter_requests(accounts_meta: List[dict], coverage_2023) -> str:
    lines = [
        "# 2023 requests (3-month chunks, max export window)",
        "",
        "Request all four quarters for each account:",
        "- 01.01.2023–31.03.2023",
        "- 01.04.2023–30.06.2023",
        "- 01.07.2023–30.09.2023",
        "- 01.10.2023–31.12.2023",
        "",
    ]
    any_needed = False
    for acc in sorted(accounts_meta, key=lambda x: x["id"]):
        alias = acc.get("alias", acc["id"])
        if acc.get("type", "").lower() != "business":
            continue
        # Determine if 2023 is already covered
        intervals = coverage_2023.get((acc["id"], alias), [])
        gaps = gaps_for_year(intervals, 2023)
        if not gaps:
            continue
        any_needed = True
        lines.append(f"- {alias} ({acc.get('type','')})")
    if not any_needed:
        lines.append("All business accounts have full 2023 coverage.")
    return "\n".join(lines) + "\n"


def digit_emoji(n: int) -> str:
    mapping = {
        0: "0️⃣",
        1: "1️⃣",
        2: "2️⃣",
        3: "3️⃣",
        4: "4️⃣",
        5: "5️⃣",
        6: "6️⃣",
        7: "7️⃣",
        8: "8️⃣",
        9: "9️⃣",
        10: "🔟",
    }
    return mapping.get(n, str(n))


def build_missing_statements(by_account, year: int) -> str:
    lines = [f"# {year} missing statements (3-month chunks)", ""]
    for (acc_id, label), intervals in by_account.items():
        gaps = gaps_for_year(intervals, year)
        chunks: List[Tuple[date, date]] = []
        for g in gaps:
            chunks.extend(chunk_gap_three_months(g))
        if not chunks:
            continue
        lines.append(f"## {display_name(acc_id, label)} — need {digit_emoji(len(chunks))} statements")
        for idx, c in enumerate(chunks, start=1):
            lines.append(f"- {idx}. {format_date(c[0])}–{format_date(c[1])}")
        lines.append("")
    content = "\n".join(lines).strip()
    if content.endswith(f"# {year} missing statements (3-month chunks)"):
        content += "\nNo missing statements for business accounts."
    return content + "\n"


def load_accounts_data(path: Path) -> List[dict]:
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8")
    data = None
    try:
        import tomllib  # type: ignore
        data = tomllib.loads(text)
    except Exception:
        try:
            import tomli  # type: ignore
            data = tomli.loads(text)
        except Exception:
            # Very simple manual parser for this specific structure
            accounts = []
            current = {}
            for line in text.splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if line.startswith("[[account]]"):
                    if current:
                        accounts.append(current)
                    current = {}
                    continue
                if "=" in line:
                    k, v = line.split("=", 1)
                    k = k.strip()
                    v = v.strip().strip('"')
                    current[k] = v
            if current:
                accounts.append(current)
            data = {"account": accounts}
    accounts = data.get("account", []) if isinstance(data, dict) else []
    # Populate global maps
    ALIAS_MAP.clear()
    TYPE_MAP.clear()
    CURRENCY_MAP.clear()
    STATUS_MAP.clear()
    for acc in accounts:
        acc_id = acc.get("id")
        if not acc_id:
            continue
        alias = acc.get("alias", acc_id)
        ALIAS_MAP[acc_id] = alias
        TYPE_MAP[acc_id] = acc.get("type", "")
        CURRENCY_MAP[acc_id] = acc.get("currency", "")
        STATUS_MAP[acc_id] = acc.get("status", "")
    return accounts


def load_statements_data(path: Path, root: Path) -> List[PeriodRange]:
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8")
    data = None
    try:
        import tomllib  # type: ignore
        data = tomllib.loads(text)
    except Exception:
        try:
            import tomli  # type: ignore
            data = tomli.loads(text)
        except Exception:
            # Minimal manual parse
            stmts = []
            current = {}
            for line in text.splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if line.startswith("[[statement]]"):
                    if current:
                        stmts.append(current)
                    current = {}
                    continue
                if "=" in line:
                    k, v = line.split("=", 1)
                    k = k.strip()
                    v = v.strip().strip('"')
                    current[k] = v
            if current:
                stmts.append(current)
            data = {"statement": stmts}
    stmts = data.get("statement", []) if isinstance(data, dict) else []
    periods: List[PeriodRange] = []
    for s in stmts:
        src = s.get("src")
        acc_id = s.get("account_id")
        alias = s.get("alias", acc_id or "")
        start = s.get("start")
        end = s.get("end")
        if not (src and acc_id and start and end):
            continue
        if not is_business(acc_id):
            continue
        try:
            start_d = datetime.fromisoformat(start).date()
            end_d = datetime.fromisoformat(end).date()
        except Exception:
            continue
        p = PeriodRange(
            path=(root / src),
            account_id=acc_id,
            account_label=alias,
            start=start_d,
            end=end_d,
            dest=(root / s["dest"]) if s.get("dest") else None,
        )
        periods.append(p)
    return periods


def main():
    root = Path(os.environ.get("DATA_ROOT", "."))
    accounts_meta = load_accounts_data(CODE_ROOT / "config" / "accounts.toml")
    # Prefer statements.toml; fallback to header scan
    periods = load_statements_data(DATA_ROOT / "metadata" / "statements.toml", DATA_ROOT)
    if not periods:
        periods = iter_periods(DATA_ROOT)
    periods.sort(key=lambda p: (p.account_id, p.start))

    with open(DATA_ROOT / "period-ranges.md", "w", encoding="utf-8") as f:
        f.write(build_period_markdown(periods))

    coverage_2024 = coverage_for_year(periods, 2024)
    with open(DATA_ROOT / "coverage-2024-table.md", "w", encoding="utf-8") as f:
        f.write(build_heatmap_table(coverage_2024, 2024))

    with open(DATA_ROOT / "tasks-2024-gaps.md", "w", encoding="utf-8") as f:
        f.write(build_gap_tasks(coverage_2024, 2024))

    with open(DATA_ROOT / "missing-2024-statements.md", "w", encoding="utf-8") as f:
        f.write(build_missing_statements(coverage_2024, 2024))

    coverage_2023 = coverage_for_year(periods, 2023)
    with open(DATA_ROOT / "tasks-2023-halfyear.md", "w", encoding="utf-8") as f:
        f.write(build_2023_quarter_requests(accounts_meta, coverage_2023))

    # Accounts alias reference
    with open(DATA_ROOT / "accounts-aliases.md", "w", encoding="utf-8") as f:
        f.write("# Account aliases (currency + type)\n\n")
        for acc in sorted(accounts_meta, key=lambda x: x["id"]):
            if acc.get("type", "").lower() != "business":
                continue
            alias = acc.get("alias", acc["id"])
            currency = acc.get("currency", "")
            acc_type = acc.get("type", "")
            status = acc.get("status", "")
            parts = [f"- {alias} \u2192 {acc['id']}", f"({currency}"]
            if acc_type:
                parts.append(acc_type)
            if status:
                parts.append(status)
            parts[-1] = parts[-1] + ")"
            f.write(" ".join(parts) + "\n")


if __name__ == "__main__":
    main()
