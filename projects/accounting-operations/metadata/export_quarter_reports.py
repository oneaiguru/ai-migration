#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
from datetime import date, datetime
import argparse
import json
import sys
from typing import Any, Dict, List

from openpyxl import load_workbook


def _cell_to_json_value(value: Any) -> Any:
    """Normalize cell value for JSON serialization."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def extract_quarter_report(path: Path) -> Dict[str, Any]:
    """Extract export amount and source row from a single quarterly XLSX report."""
    wb = load_workbook(path, data_only=True)
    try:
        if "Лист1" in wb.sheetnames:
            ws = wb["Лист1"]
        else:
            ws = wb.worksheets[0]

        header_row_idx = 10
        data_row_idx = 11

        # Build source row object from A11..H11 with headers from row 10.
        cells: Dict[str, Dict[str, Any]] = {}
        revenue_value: Any = None
        for col_idx in range(1, 9):  # columns A..H
            header_cell = ws.cell(row=header_row_idx, column=col_idx)
            data_cell = ws.cell(row=data_row_idx, column=col_idx)
            header = header_cell.value
            value = _cell_to_json_value(data_cell.value)
            col_letter = data_cell.column_letter
            if col_letter == "F":
                revenue_value = data_cell.value
            cells[col_letter] = {
                "cell": f"{col_letter}{data_row_idx}",
                "header": header,
                "value": value,
            }

        if revenue_value is None:
            revenue_value = 0

        export_label_cell = ws["D17"]
        export_amount_cell = ws["F17"]

        result: Dict[str, Any] = {
            "file": path.name,
            "sheet": ws.title,
            "export": {
                "label_cell": export_label_cell.coordinate,
                "label": export_label_cell.value,
                "amount_cell": export_amount_cell.coordinate,
                "amount": _cell_to_json_value(revenue_value),
            },
            "source_row": {
                "index": data_row_idx,
                "cells": cells,
            },
        }

        return result
    finally:
        wb.close()


def collect_export_rows(reports_root: Path) -> List[Dict[str, Any]]:
    """Collect export rows from all quarterly reports in a folder."""
    if not reports_root.exists() or not reports_root.is_dir():
        raise FileNotFoundError(f"Reports root does not exist or is not a directory: {reports_root}")

    entries: List[Dict[str, Any]] = []
    for path in sorted(reports_root.glob("*.xlsx")):
        entries.append(extract_quarter_report(path))
    return entries


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract export rows from quarterly XLSX reports into JSON."
    )
    parser.add_argument(
        "--reports-root",
        type=str,
        required=True,
        help="Folder that contains quarterly reports named 'YYYY_QN.xlsx'.",
    )
    parser.add_argument(
        "--out",
        type=str,
        required=True,
        help="Path to the JSON file to write.",
    )
    return parser.parse_args(argv)


def main(argv: List[str] | None = None) -> None:
    args = parse_args(argv or sys.argv[1:])
    reports_root = Path(args.reports_root).expanduser().resolve()
    out_path = Path(args.out).expanduser()

    entries = collect_export_rows(reports_root)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {out_path} with {len(entries)} entries")


if __name__ == "__main__":
    main()

