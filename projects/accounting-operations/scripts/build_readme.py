#!/usr/bin/env python3
from __future__ import annotations

from datetime import datetime, date, timedelta
from pathlib import Path
import os
from typing import Optional, List, Tuple
import sys

from openpyxl import load_workbook

# Paths
SCRIPT_ROOT = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_ROOT.parent
DATA_ROOT = Path(os.environ.get("DATA_ROOT", REPO_ROOT)).resolve()
OUTPUT_ROOT = Path(os.environ.get("OUTPUT_ROOT", REPO_ROOT / "docs" / "current")).resolve()

# Ensure repo root on path for imports
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from generate_reports import load_accounts_data, load_statements_data, display_name  # noqa: E402


def find_header_row(ws) -> Optional[int]:
    for r in range(1, 15):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if any(isinstance(v, str) and "Дата" in v for v in row_vals):
            return r
    return None


def parse_date(val) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except Exception:
                continue
    return None


def has_transactions(path: Path) -> bool:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header_row = find_header_row(ws)
    if header_row is None:
        return False
    for r in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if parse_date(val):
            return True
    return False


def merge_gaps(intervals: List[Tuple[date, date]]) -> List[Tuple[date, date]]:
    intervals = sorted(intervals, key=lambda x: x[0])
    merged = []
    for s, e in intervals:
        if not merged or s > merged[-1][1] + timedelta(days=1):
            merged.append((s, e))
        else:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))
    return merged


def find_gaps(intervals: List[Tuple[date, date]]) -> List[Tuple[date, date]]:
    if not intervals:
        return []
    ints = merge_gaps(intervals)
    gaps = []
    for i in range(len(ints) - 1):
        prev_end = ints[i][1]
        next_start = ints[i + 1][0]
        if next_start > prev_end + timedelta(days=1):
            gaps.append((prev_end + timedelta(days=1), next_start - timedelta(days=1)))
    return gaps


def main():
    load_accounts_data(REPO_ROOT / "config" / "accounts.toml")
    stmts = load_statements_data(DATA_ROOT / "metadata" / "statements.toml", DATA_ROOT)
    by_account = {}
    for s in stmts:
        by_account.setdefault(s.account_id, []).append(s)

    lines = ["# Пакет выписок (2023–2024) — только бизнес-счета", ""]
    lines.append("Файлы лежат в `renamed-statements/`. Статусы: ✅ есть движения; ⚠️ нет движений.")
    lines.append("")

    for acc_id, statements in sorted(by_account.items()):
        alias = display_name(acc_id, statements[0].account_label)
        lines.append(f"## {alias}")
        intervals = [(s.start, s.end) for s in sorted(statements, key=lambda x: x.start)]
        gaps = find_gaps(intervals)
        total_files = len(statements)
        if gaps:
            gap_text = "; ".join(f"{g[0].strftime('%d.%m.%Y')}–{g[1].strftime('%d.%m.%Y')}" for g in gaps)
            lines.append(f"- Файлов: {total_files}; Обнаружены разрывы: {gap_text}")
        else:
            lines.append(f"- Файлов: {total_files}; Без разрывов по датам")
        lines.append("")
        lines.append("| Файл | Период | Статус |")
        lines.append("| --- | --- | --- |")
        for s in sorted(statements, key=lambda x: x.start):
            dest_path = DATA_ROOT / (s.dest or s.path)
            dest_rel = dest_path.relative_to(DATA_ROOT).as_posix()
            status = "✅" if has_transactions(dest_path) else "⚠️ (нет движений)"
            period = f"{s.start.strftime('%d.%m.%Y')}–{s.end.strftime('%d.%m.%Y')}"
            lines.append(f"| `{dest_rel}` | {period} | {status} |")
        lines.append("")

    out_dir = OUTPUT_ROOT
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "README.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
