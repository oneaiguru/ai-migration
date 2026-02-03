import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from metadata.export_quarter_reports import collect_export_rows


class ExportQuarterReportsTest(unittest.TestCase):
    def _create_quarter_workbook(self, path: Path, quarter_label: str, export_amount: float) -> None:
        """Create a minimal quarterly workbook matching the PVT template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"

        # Header row (row 10)
        headers = [
            "Месяц, год",
            "Расчетный квартал",
            "Дата признания дохода",
            "Вид деятельности",
            "Заказчик услуги ",
            "Выручка от реализации услуг, сом",
            "Сумма отчислений  в дирекцию ПВТ (сом)",
            "Количество работников (чел.)",
        ]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=10, column=idx, value=header)

        # Data row (row 11)
        ws["A11"] = "2023-01-01"
        ws["B11"] = quarter_label
        ws["C11"] = None
        ws["D11"] = "Some activity"
        ws["E11"] = "Some customer"
        ws["F11"] = export_amount
        ws["G11"] = export_amount * 0.01
        ws["H11"] = 0

        # Export row (row 17)
        ws["D17"] = "Экспорт"
        ws["F17"] = "=F11"

        wb.save(path)
        wb.close()

    def test_collect_export_rows_basic(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)

            # Create three quarter reports
            self._create_quarter_workbook(root / "2023_Q1.xlsx", "1 квартал 2023 года", 100.0)
            self._create_quarter_workbook(root / "2023_Q3.xlsx", "3 квартал 2023 года", 200.0)
            self._create_quarter_workbook(root / "2024_Q4.xlsx", "4 квартал 2024 года", 300.0)

            entries = collect_export_rows(root)

            self.assertEqual(len(entries), 3)
            files = {e["file"] for e in entries}
            self.assertEqual(files, {"2023_Q1.xlsx", "2023_Q3.xlsx", "2024_Q4.xlsx"})

            for entry in entries:
                export = entry["export"]
                source_row = entry["source_row"]
                cells = source_row["cells"]

                # We always expect the label to be "Экспорт"
                self.assertEqual(export["label"], "Экспорт")
                self.assertEqual(export["amount_cell"], "F17")
                self.assertEqual(source_row["index"], 11)

                # F cell in the source row must match the export amount
                self.assertIn("F", cells)
                self.assertEqual(cells["F"]["cell"], "F11")
                self.assertEqual(cells["F"]["value"], export["amount"])


if __name__ == "__main__":
    unittest.main()

