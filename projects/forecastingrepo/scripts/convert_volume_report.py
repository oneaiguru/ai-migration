#!/usr/bin/env python3
"""
Ingest "Отчет по объёмам ..." reports (XLSX or CSV) and emit canonical site service data.

Outputs:
  * data/sites/sites_service.csv   — site_id, service_dt, collect_volume_m3, volume_m3, source
  * data/sites/sites_registry.csv  — site_id, district, address, source

Duplicate site/day rows are summed. The parser handles both CSV exports and the original XLSX workbooks by
reading sheet XML via openpyxl in read-only mode.
"""
from __future__ import annotations

import argparse
import calendar
import csv
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


DATE_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})")


@dataclass
class DailyRecord:
    site_id: str
    service_dt: date
    volume_m3: float
    collect_volume_m3: float
    source: str


@dataclass
class RegistryRecord:
    site_id: str
    district: Optional[str]
    address: Optional[str]
    source: str


def parse_period(text: str) -> Optional[Tuple[date, date]]:
    m = DATE_RE.search(text)
    if not m:
        return None
    start = datetime.strptime(m.group(1), "%d.%m.%Y").date()
    end = datetime.strptime(m.group(2), "%d.%m.%Y").date()
    return start, end


def parse_float(value: str) -> Optional[float]:
    s = value.strip()
    if not s or s in {"-", "—"}:
        return None
    s = s.replace(" ", "")
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def first_positive_number(cell: str | None) -> Optional[float]:
    if not cell:
        return None
    for token in re.split(r"[\\/|;]", cell):
        val = parse_float(token)
        if val and val > 0:
            return val
    return None


def iterate_csv_rows(path: Path) -> Iterator[List[str]]:
    encodings = ("utf-8", "cp1251")
    last_exc = None
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, errors="ignore") as f:
                sample = f.read(1024)
                f.seek(0)
                delimiter = ";" if sample.count(";") >= sample.count(",") else ","
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    yield [cell.strip() for cell in row]
            return
        except Exception as exc:  # pragma: no cover
            last_exc = exc
            continue
    raise RuntimeError(f"Unable to read CSV {path}: {last_exc}")  # pragma: no cover


def iterate_xlsx_rows(path: Path) -> Iterator[List[str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read XLSX files")
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            values = []
            for cell in row:
                if cell is None:
                    values.append("")
                elif isinstance(cell, datetime):
                    values.append(cell.strftime("%d.%m.%Y"))
                elif isinstance(cell, (int, float)):
                    values.append(str(cell))
                else:
                    values.append(str(cell))
            if any(v.strip() for v in values):
                yield [v.strip() for v in values]
    wb.close()


def read_rows(path: Path) -> Tuple[Tuple[date, date], List[str], List[List[str]]]:
    row_iter: Iterator[List[str]]
    ext = path.suffix.lower()
    if ext in {".csv", ".txt"}:
        row_iter = iterate_csv_rows(path)
    elif ext in {".xlsx", ".xlsm"}:
        row_iter = iterate_xlsx_rows(path)
    else:
        raise RuntimeError(f"Unsupported file type: {path}")

    period: Optional[Tuple[date, date]] = None
    header: Optional[List[str]] = None
    data_rows: List[List[str]] = []
    for raw_row in row_iter:
        if not any(raw_row):
            continue
        joined = " ".join(raw_row)
        if not period:
            maybe = parse_period(joined)
            if maybe:
                period = maybe
        if header is None and raw_row[0].startswith("N"):
            header = raw_row
            continue
        if header:
            data_rows.append(raw_row)
    if period is None:
        raise RuntimeError(f"Period line not found in {path}")
    if header is None:
        raise RuntimeError(f"Header row not found in {path}")
    return period, header, data_rows


def read_records(path: Path) -> Tuple[List[DailyRecord], List[RegistryRecord]]:
    period, header_raw, data_rows = read_rows(path)
    period_start, period_end = period

    header = [cell.strip() for cell in header_raw]

    def find_index(keys: Iterable[str], header_lower: List[str]) -> int:
        for idx, name in enumerate(header_lower):
            if name in keys:
                return idx
        return -1

    header_lower = [name.strip().lower() for name in header]
    idx_site = find_index(
        {"код кп", "код кп.", "кп", "код площадки", "код контейнерной площадки", "код kp", "код"} , header_lower
    )
    idx_total = find_index({"итого, м3", "итого, м^3", "итого, м³", "итог, м3"}, header_lower)
    idx_graph = find_index({"график вывоза"}, header_lower)
    idx_district = find_index({"район", "участок"}, header_lower)
    idx_address = find_index({"адрес", "адрес кп"}, header_lower)
    idx_count = find_index({"количество", "кол-во"}, header_lower)
    idx_volume_meta = find_index({"объем", "объём", "объем, м3", "объём, м3"}, header_lower)

    if idx_site < 0 or idx_graph < 0 or idx_total < 0:
        raise RuntimeError(f"Required columns not found in {path}")

    daily_start = idx_graph + 1
    daily_end = idx_total
    month_blocks: List[List[int]] = []
    current: List[int] = []
    for col_idx in range(daily_start, daily_end):
        if col_idx >= len(header):
            break
        cell = header[col_idx]
        cell_clean = cell.strip()
        try:
            day_num = int(cell_clean)
        except ValueError:
            day_num = None
        if day_num == 1 and current:
            month_blocks.append(current)
            current = [col_idx]
        else:
            current.append(col_idx)
    if current:
        month_blocks.append(current)

    months: List[Tuple[int, int]] = []
    year, month = period_start.year, period_start.month
    while True:
        months.append((year, month))
        if (year, month) == (period_end.year, period_end.month):
            break
        month += 1
        if month > 12:
            month = 1
            year += 1

    if len(month_blocks) != len(months):
        raise RuntimeError(
            f"Detected {len(month_blocks)} month column blocks but period spans {len(months)} months in {path}"
        )

    daily_records: List[DailyRecord] = []
    registry_records: Dict[str, RegistryRecord] = {}

    for row in data_rows:
        if idx_site >= len(row):
            continue
        site_id = row[idx_site].strip()
        if not site_id:
            continue

        if site_id not in registry_records:
            district = row[idx_district].strip() if 0 <= idx_district < len(row) and row[idx_district] else None
            address = row[idx_address].strip() if 0 <= idx_address < len(row) and row[idx_address] else None
            registry_records[site_id] = RegistryRecord(
                site_id=site_id,
                district=district,
                address=address,
                source=path.name,
            )

        total_volume = parse_float(row[idx_total]) if idx_total < len(row) else None
        for month_idx, (year, month) in enumerate(months):
            columns = month_blocks[month_idx]
            last_day = calendar.monthrange(year, month)[1]
            for col_idx in columns:
                if col_idx >= len(row):
                    continue
                header_val = header[col_idx]
                try:
                    day = int(header_val.strip())
                except ValueError:
                    continue
                if day > last_day:
                    continue
                try:
                    value = parse_float(row[col_idx])
                except Exception:
                    value = None
                if value is None or value <= 0:
                    continue
                service_dt = date(year, month, day)
                if service_dt < period_start or service_dt > period_end:
                    continue
                daily_records.append(
                    DailyRecord(
                        site_id=site_id,
                        service_dt=service_dt,
                        volume_m3=value,
                        collect_volume_m3=value,
                        source=path.name,
                    )
                )

    return daily_records, list(registry_records.values())


def merge_daily(existing_path: Path, new_records: List[DailyRecord]) -> None:
    combined: Dict[Tuple[str, date], DailyRecord] = {}
    if existing_path.exists():
        with existing_path.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for rec in reader:
                site_id = rec["site_id"]
                dt = datetime.strptime(rec["service_dt"], "%Y-%m-%d").date()
                combined[(site_id, dt)] = DailyRecord(
                    site_id=site_id,
                    service_dt=dt,
                    volume_m3=parse_float(rec.get("volume_m3") or "") or 0.0,
                    collect_volume_m3=parse_float(rec.get("collect_volume_m3") or "") or 0.0,
                    source=rec.get("source", ""),
                )

    for rec in new_records:
        key = (rec.site_id, rec.service_dt)
        if key in combined:
            existing = combined[key]
            existing.volume_m3 += rec.volume_m3
            existing.collect_volume_m3 += rec.collect_volume_m3
            sources = set(filter(None, existing.source.split(";")))
            sources.add(rec.source)
            existing.source = ";".join(sorted(sources))
        else:
            combined[key] = rec

    with existing_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["site_id", "service_dt", "collect_volume_m3", "volume_m3", "source"])
        for key in sorted(combined.keys()):
            rec = combined[key]
            writer.writerow(
                [
                    rec.site_id,
                    rec.service_dt.isoformat(),
                    f"{rec.collect_volume_m3:.6f}",
                    f"{rec.volume_m3:.6f}",
                    rec.source,
                ]
            )


def merge_registry(existing_path: Path, new_records: List[RegistryRecord]) -> None:
    combined: Dict[str, RegistryRecord] = {}
    if existing_path.exists():
        with existing_path.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for rec in reader:
                combined[rec["site_id"]] = RegistryRecord(
                    site_id=rec["site_id"],
                    district=rec.get("district") or None,
                    address=rec.get("address") or None,
                    source=rec.get("source", ""),
                )

    for rec in new_records:
        existing = combined.get(rec.site_id)
        if existing:
            if not existing.district and rec.district:
                existing.district = rec.district
            if not existing.address and rec.address:
                existing.address = rec.address
            sources = set(filter(None, existing.source.split(";")))
            if rec.source:
                sources.add(rec.source)
            existing.source = ";".join(sorted(sources))
        else:
            combined[rec.site_id] = rec

    with existing_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["site_id", "district", "address", "source"])
        for site_id in sorted(combined.keys()):
            rec = combined[site_id]
            writer.writerow([rec.site_id, rec.district or "", rec.address or "", rec.source])


def summarize(records: List[DailyRecord]) -> str:
    total_rows = len(records)
    total_sites = len({rec.site_id for rec in records})
    return f"[OK] Parsed {total_rows} daily rows across {total_sites} sites"


def main(argv: Optional[Sequence[str]] = None) -> None:
    parser = argparse.ArgumentParser(description="Convert volume reports (CSV/XLSX) into canonical site service CSV.")
    parser.add_argument("--input", nargs="+", required=True, help="Paths to Отчет по объёмам files (.xlsx or .csv)")
    parser.add_argument("--output-service", default="data/sites/sites_service.csv")
    parser.add_argument("--output-registry", default="data/sites/sites_registry.csv")
    args = parser.parse_args(argv)

    out_service = Path(args.output_service)
    out_registry = Path(args.output_registry)
    out_service.parent.mkdir(parents=True, exist_ok=True)
    out_registry.parent.mkdir(parents=True, exist_ok=True)

    all_daily: List[DailyRecord] = []
    all_registry: List[RegistryRecord] = []
    for input_path in args.input:
        path = Path(input_path)
        daily_records, registry_records = read_records(path)
        all_daily.extend(daily_records)
        all_registry.extend(registry_records)
        print(summarize(daily_records))

    merge_daily(out_service, all_daily)
    merge_registry(out_registry, all_registry)
    print(f"[DONE] Wrote {out_service} and {out_registry}")


if __name__ == "__main__":
    main()
