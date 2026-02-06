#!/usr/bin/env python3
"""
Extract per-site daily actuals from "Отчет по объемам ... для всех участков" Excel files.

The Excel sheet structure (observed):
  - Header rows include a line like: "Период: dd.mm.yyyy - dd.mm.yyyy"
  - A table with columns such as: N п/п, Район, Участок, Лот, Код КП, Адрес, ...
  - Followed by repeated day-of-month columns labeled 1..31 many times (one block per month in the report period).

This script unpivots the daily columns into a long CSV with:
  site_id, service_dt, collect_volume_m3

Notes:
  - We treat values in day columns as volume in m^3.
  - The report may span multiple months (e.g., H1/H2). Use --month to filter a single YYYY-MM.
  - We do not commit XLSX to the repo. Use this tool to create tidy CSVs outside git.
"""
from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook  # type: ignore
except Exception as e:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from e


@dataclass
class HeaderLayout:
    header_row: int
    site_id_col: int
    start_date: date
    end_date: date
    day_blocks: List[Tuple[date, List[int]]]


def _parse_period(ws) -> Optional[Tuple[date, date]]:
    # Scan top 10 rows for a line containing "Период: dd.mm.yyyy - dd.mm.yyyy"
    rx = re.compile(r"Период\s*:\s*(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        text = " ".join(str(c) for c in row if c)
        m = rx.search(text)
        if m:
            d1 = datetime.strptime(m.group(1), "%d.%m.%Y").date()
            d2 = datetime.strptime(m.group(2), "%d.%m.%Y").date()
            return d1, d2
    return None


def _days_in_month(dt: date) -> int:
    from calendar import monthrange
    return monthrange(dt.year, dt.month)[1]


def detect_layout(xlsx_path: Path) -> HeaderLayout:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 1) Period
    period = _parse_period(ws)
    if not period:
        wb.close()
        raise SystemExit("Could not find 'Период: dd.mm.yyyy - dd.mm.yyyy' in the top rows")
    start_date, end_date = period

    # 2) Find header row (contains 'Код КП')
    header_row = None
    site_id_col = None
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any("Код КП" in v for v in vals):
            header_row = i
            headers = vals
            for idx, name in enumerate(vals, start=1):
                if name.strip() == "Код КП":
                    site_id_col = idx
                    break
            break
    if header_row is None or site_id_col is None:
        wb.close()
        raise SystemExit("Could not locate header row or 'Код КП' column")

    # 3) Locate useful column indices and first day column index
    district_col = None
    for idx, name in enumerate(headers, start=1):
        if name.strip() == "Район":
            district_col = idx
            break
    # first exact header that is an integer 1..31 after the fixed columns
    day_header_indices: List[int] = []
    for idx, name in enumerate(headers, start=1):
        if name and name.isdigit():
            day_header_indices.append(idx)
    if not day_header_indices:
        wb.close()
        raise SystemExit("No day-of-month columns (1..31) detected in header")

    first_day_idx = min(day_header_indices)

    # 4) Build blocks of day columns per month from start_date to end_date.
    blocks: List[Tuple[date, List[int]]] = []
    cur = date(start_date.year, start_date.month, 1)
    pos = first_day_idx
    # Get total headers length for boundary
    hdr_len = len(headers)
    while cur <= end_date and pos <= hdr_len:
        dim = _days_in_month(cur)
        cols = list(range(pos, min(pos + dim, hdr_len + 1)))
        blocks.append((cur, cols))
        pos += dim
        # advance month
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)

    wb.close()
    layout = HeaderLayout(header_row=header_row, site_id_col=site_id_col, start_date=start_date, end_date=end_date, day_blocks=blocks)
    # attach dynamically for downstream read (keep dataclass minimal for compatibility)
    setattr(layout, "district_col", district_col)
    setattr(layout, "headers", headers)
    return layout


def _to_float(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(" ", "")
    # tolerate decimal comma
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    # strip any annotations like "0.75 / 0.75"
    if "/" in s:
        s = s.split("/")[0].strip()
    try:
        return float(s)
    except Exception:
        return None


def extract_csv(xlsx_path: Path, out_csv: Path, month_filter: Optional[str] = None) -> int:
    """Return number of rows written."""
    layout = detect_layout(xlsx_path)
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Parse month filter
    mf_year = mf_month = None
    if month_filter:
        try:
            dt = datetime.strptime(month_filter, "%Y-%m").date()
            mf_year, mf_month = dt.year, dt.month
        except Exception:
            wb.close()
            raise SystemExit("--month must be YYYY-MM")

    # Build list of (month_start_date, day_col_indices) we will output
    blocks = layout.day_blocks
    if mf_year and mf_month:
        blocks = [b for b in blocks if b[0].year == mf_year and b[0].month == mf_month]

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    written = 0
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        # Include district if present
        include_district = getattr(layout, "district_col", None) is not None
        header = ["site_id", "service_dt", "collect_volume_m3"]
        if include_district:
            header.insert(1, "district")
        w.writerow(header)  # unit: m^3
        # Iterate rows beneath header
        for row in ws.iter_rows(min_row=layout.header_row + 1, values_only=True):
            site_id_val = row[layout.site_id_col - 1] if len(row) >= layout.site_id_col else None
            site_id = str(site_id_val).strip() if site_id_val not in (None, "") else ""
            if not site_id:
                continue
            district_val = None
            if include_district:
                dc = getattr(layout, "district_col", None)
                if dc and len(row) >= dc:
                    val = row[dc - 1]
                    district_val = str(val).strip() if val not in (None, "") else ""
            for month_start, cols in blocks:
                for day_offset, col_idx in enumerate(cols, start=1):
                    # date within month
                    try:
                        d = date(month_start.year, month_start.month, day_offset)
                    except Exception:
                        continue
                    # column index is 1-based
                    val = row[col_idx - 1] if len(row) >= col_idx else None
                    v = _to_float(val)
                    if v is None:
                        continue
                    rec = [site_id, d.isoformat(), f"{v:.6f}"]
                    if include_district:
                        rec = [site_id, (district_val or ""), d.isoformat(), f"{v:.6f}"]
                    w.writerow(rec)
                    written += 1
    wb.close()
    return written


def main(argv: List[str] | None = None) -> None:
    ap = argparse.ArgumentParser(description="Unpivot 'Отчет по объемам' Excel to CSV of per-site daily volume")
    ap.add_argument("--xlsx", required=True, help="Path to Excel 'Отчет по объемам' file")
    ap.add_argument("--out", required=True, help="Output CSV path")
    ap.add_argument("--month", help="Optional month filter YYYY-MM (e.g., 2024-08)")
    args = ap.parse_args(argv)

    xlsx = Path(args.xlsx)
    out = Path(args.out)
    n = extract_csv(xlsx, out, month_filter=args.month)
    print(f"[OK] Wrote {n} rows to {out}")


if __name__ == "__main__":  # pragma: no cover
    main()
